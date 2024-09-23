# """
# Developed by Azhagan Avr (aavr@kittelson.com)
# V4
# Upgraded code to work with Streamlit
# """

# Libraries
import pandas as pd # pandas, python-dateutil, pytz
import numpy as np
import math
import time
import logging
import datetime as dt
import sys
import streamlit as st
from io import BytesIO
import requests
from openpyxl import load_workbook
import zipfile
import io

# logging.getLogger('PIL').setLevel(# logging.WARNING)

# Initialize session state to manage download state
if 'downloaded' not in st.session_state:
    st.session_state.downloaded = False


def get_file_format(start_value):
    return st.selectbox("Select the file format", [start_value,'txt', 'csv'])


def get_uploaded_files(file):
    return st.file_uploader(f"Upload {file} files", accept_multiple_files=True, type=[file])


def get_dict_mod_raw_data(raw_df):

    dict_df = pd.read_excel(get_url("dict"))

    temp = dict_df.drop(columns='Values')  # droping the values column
    cols = temp.columns.tolist()
    raw_df.columns = cols    
    return dict_df, raw_df


def get_output_file_name():
    return st.text_input("Please enter the output filename: ")


def concatenate_files(uploaded_files, format_file):
    if format_file == 'txt':
        content = ""
        for uploaded_file in uploaded_files:
            content += uploaded_file.getvalue().decode("utf-8").strip() + "\n"
        content = content.strip()
        lines = content.splitlines()
        data = [line.split(',') for line in lines]
        concatenated_df = pd.DataFrame(data)
    elif format_file == 'csv':
        dfs = []
        for uploaded_file in uploaded_files:
            df = pd.read_csv(uploaded_file)
            dfs.append(df)
        concatenated_df = pd.concat(dfs, ignore_index=True)
    else:
        raise ValueError("Unsupported file format. Please select either 'txt' or 'csv'.")
    
    # Remove completely empty rows
    concatenated_df.dropna(how='all', inplace=True)
    return concatenated_df


def data_translation(raw_df, translation_df, start_time):

    try:
        status_text = st.empty()
        index_values_float = translation_df["Values"].tolist()  # converting values to a list and it is considered to be a float
        index_values_str = [str(idx) for idx in index_values_float]  # list is converted to string
        iterate = 0
        no_of_columns = len(raw_df.columns.tolist())
        for df_col in raw_df.columns:
            iterate = iterate+1
            unique_elements = translation_df[df_col].unique() # getting the unique values in the dictionary spreadsheet for the chosen column
            status_text.text("\r"+str(round(time.time() - start_time, ndigits=0))+"s: Translating Column >> " + str(iterate)+"/"+str(no_of_columns))
            # replacing the values of columns that are float or int
            if str(raw_df[df_col].dtype) == "float64" or str(raw_df[df_col].dtype) == "int64":
                if len(unique_elements) > 1:
                    for i in range(len(raw_df[df_col])):
                        if not math.isnan(raw_df[df_col][i]):
                           raw_df.loc[i,df_col] = translation_df[df_col][index_values_float.index((raw_df[df_col][i]))]
                        else:
                            pass
                else:
                    # st.text("No data to translate")
                    pass

            # replacing the values of columns that are not float or int and considered to be object type
            else:
                # st.text(df_col, raw_df[df_col].dtype)
                # st.text(df_col)
                if len(unique_elements) > 1:
                    for i in range(len(raw_df[df_col])):
                        if len(str(raw_df[df_col][i])) > 1:
                            if str(raw_df[df_col][i]) != "00":
                                try:
                                    # st.text(len((raw_df[df_col][i])))
                                    raw_df.loc[i, df_col] = translation_df[df_col][index_values_str.index(str(raw_df[df_col][i]).lstrip("0"))]
                                except Exception as err:
                                    # st.text("hi")
                                    pass
                            else:
                                try:
                                    # st.text(len((raw_df[df_col][i])))
                                    raw_df.loc[i, df_col] = translation_df[df_col][index_values_str.index(str(0))]
                                except Exception as err:
                                    # st.text("hi")
                                    pass
                        else:
                            try:
                                # st.text(len((raw_df[df_col][i])))
                                raw_df.loc[i, df_col] = translation_df[df_col][index_values_str.index(str(raw_df[df_col][i]))]
                            except Exception as err:
                                # st.text("hi")
                                pass
                else:
                    pass
                    # st.text("No data to translate")

        raw_df["Latitude Degrees"] = pd.to_numeric(raw_df['Latitude Degrees'], errors='coerce').astype(float)
        raw_df["Latitude Minutes"] = pd.to_numeric(raw_df['Latitude Minutes'], errors='coerce').astype(float)
        raw_df["Latitude Seconds"] = pd.to_numeric(raw_df['Latitude Seconds'], errors='coerce').astype(float)
        raw_df["Longitude Degrees"] = pd.to_numeric(raw_df['Longitude Degrees'], errors='coerce').astype(float)
        raw_df["Longitude Minutes"] = pd.to_numeric(raw_df['Longitude Minutes'], errors='coerce').astype(float)
        raw_df["Longitude Seconds"] = pd.to_numeric(raw_df['Longitude Seconds'], errors='coerce').astype(float)
        latitude = raw_df["Latitude Degrees"]+(raw_df["Latitude Minutes"]/60)+(raw_df["Latitude Seconds"]/3600)
        longitude = raw_df["Longitude Degrees"] - ((raw_df["Longitude Minutes"] / 60) + (raw_df["Longitude Seconds"] / 3600))

        raw_df.insert(raw_df.columns.get_loc('Longitude Seconds')+1,"Latitude", latitude)
        raw_df.insert(raw_df.columns.get_loc('Longitude Seconds')+2, "Longitude", longitude)
        # raw_df.to_csv(path_dir+"Translated_Data.csv",index=False)
        return (raw_df)

    except Exception as err:
        # logging.error(err)
        st.text(err)
        sys.exit(1)


def participant_vehicle_id(raw_df, veh_code):
    # creating ids for each participant in a crash
    try:
        # raw_df = pd.read_csv(path_dir + file_name, low_memory=False)
        raw_df.rename(columns={'Crash ID': 'ID'}, inplace=True)
        participant_vehicle_id_lst = []

        numeric_columns = ["Record Type", "Vehicle ID"]
        for col in numeric_columns:
            raw_df[col] = pd.to_numeric(raw_df[col], errors='coerce')


        for idx in range(len(raw_df["ID"])):
            if raw_df["Record Type"][idx] == 1:  # if record type  is 1, then the crash record is general crash info
                participant_vehicle_id_lst.append("General Crash Information")

            elif raw_df["Record Type"][idx] == 2:  # if record type  is 2, then the crash record is denoting the vehicle data
                if veh_code == "1":
                    participant_vehicle_id_lst.append("Vehicle "+str(int(raw_df["Participant Vehicle Seq#"][idx])))  # depreceated
                else:
                    participant_vehicle_id_lst.append("Vehicle " + str(int(raw_df["Vehicle Coded Seq#"][idx])))

            elif raw_df["Record Type"][idx] == 3:  # if record type  is 3, then the crash record is denoting the Participant data
                if raw_df["Vehicle ID"][idx] != 0:  # Participant belonging to a vehicle
                    participant_vehicle_id_lst.append("Vehicle "+str(str(int(raw_df["Vehicle Coded Seq#"][idx]))+" Participant "+str(int(raw_df["Participant Vehicle Seq#"][idx]))))
                else:  # then the data denotes a pedestrian
                    participant_vehicle_id_lst.append("Pedestrian/Pedalcyclist "+str(int(raw_df["Participant Vehicle Seq#"][idx])))
            else:
                pass

        raw_df.insert(1, "Participant Vehicle ID", participant_vehicle_id_lst)
        # raw_df.to_csv(path_dir + "Translated_Data.csv", index=False)

        return raw_df
    except Exception as err:
        st.text(err)
        # logging.error(err)
        sys.exit(1)


def add_kai_variables(raw_df):
    try:
        raw_df = raw_df.drop_duplicates().reset_index()
        k01_sevr = []
        k02_sevrcl = []
        k03_sevrfl = []
        k04_epdo = []
        k06_imprfl = []
        k07_aggfl = []
        k08_rdepfl = []
        k09_wildfl = []
        k10_disdfl = []
        k11_intxfl = []
        k13_pedfl = []
        k14_pcycfl = []
        k62_trnfl = []
        k63_protfl = []
        k65_peddi = []
        k66_pcycdi = []
        
        numeric_columns = ["ID","Record Type", "Vehicle ID", "Crash Year","Total Fatality Count","Total Suspected Serious Injury (A) Count",
                           "Total Suspected Minor Injury (B) Count","Total Possible Injury (C) Count", "Total Pedestrian Count", "Total Pedalcyclist Count",
                           "Age"]
        for col in numeric_columns:
            raw_df[col] = pd.to_numeric(raw_df[col], errors='coerce')


        # Filter Definitions: refer to the definition spreadsheet document OTSDESpecificCrashFlags/50000_recoder mapping for new variablesc
        k08_rdepfl_filter_1 = ["Street/road or highway intersection", "Turning Movement", "passing situation",
                               "Motor vehicle on other roadway", "Sideswipe-meeting",
                               "From opposite direction - both going straight", "Head-On", "Wrong way on one-way roadway", "Fixed object"]
        k08_rdepfl_filter_2 = ["Straddling or driving on wrong lanes", "Failed to negotiate a curve",
                               "Failed to maintain lane", "Ran off road"]
        k08_rdepfl_filter_3 = ["Bridge girder (horizontal structure overhead)", "Tree branch or other vegetation overhead, etc.",
                               "Wire or cable across or over the road", "Slides, rocks off or on road, falling rocks",
                               "Speed bump, other bump, pothole or pavement irregularity (Per PAR) (2014)",
                               "Other overhead object (highway sign, signal head, etc.); not bridge",
                               "Expansion joint", "Rock slide or land slide"]
        k08_rdepfl_filter_4 = raw_df[(raw_df["Road Character"] != k08_rdepfl_filter_1[0]) &
                                     (raw_df["Intersection Related Flag"] != "Yes") &
                                     (raw_df["Collision Type"] != k08_rdepfl_filter_1[1]) &
                                     (raw_df['Participant Type Code'] == 'Driver') &
                                     (raw_df['Vehicle Action Code'] != k08_rdepfl_filter_1[2]) &
                                     (
                                             (raw_df["Off Roadway Flag"] == "Yes") |
                                             (raw_df["Crash Type"] == k08_rdepfl_filter_1[3]) |
                                             (raw_df["Collision Type"] == k08_rdepfl_filter_1[4]) |
                                             (
                                                     (raw_df["Crash Type"] == k08_rdepfl_filter_1[5]) &
                                                     (raw_df["Collision Type"] == k08_rdepfl_filter_1[6]) &
                                                     (raw_df['Crash Level Cause 1 Code'] == k08_rdepfl_filter_1[7]) &
                                                     (raw_df['Crash Level Cause 2 Code'] == k08_rdepfl_filter_1[7]) &
                                                     (raw_df['Crash Level Cause 3 Code'] == k08_rdepfl_filter_1[7]) &
                                                     (raw_df['Participant Cause 1 Code'] == k08_rdepfl_filter_1[7]) &
                                                     (raw_df['Participant Cause 2 Code'] == k08_rdepfl_filter_1[7]) &
                                                     (raw_df['Participant Cause 3 Code'] == k08_rdepfl_filter_1[7])
                                             ) |

                                             (
                                                     (raw_df['Participant Error 1 Code'].isin(k08_rdepfl_filter_2)) |
                                                     (raw_df['Participant Error 2 Code'].isin(k08_rdepfl_filter_2)) |
                                                     (raw_df['Participant Error 3 Code'].isin(k08_rdepfl_filter_2))

                                             ) |

                                             (
                                                     (raw_df["Crash Type"] == k08_rdepfl_filter_1[8]) &
                                                     (~raw_df["Vehicle Event 1 Code"].isin(k08_rdepfl_filter_3)) &
                                                     (~raw_df["Vehicle Event 2 Code"].isin(k08_rdepfl_filter_3)) &
                                                     (~raw_df["Vehicle Event 3 Code"].isin(k08_rdepfl_filter_3)) &
                                                     (~raw_df["Crash Level Event 1 Code"].isin(k08_rdepfl_filter_3)) &
                                                     (~raw_df["Crash Level Event 2 Code"].isin(k08_rdepfl_filter_3)) &
                                                     (~raw_df["Crash Level Event 3 Code"].isin(k08_rdepfl_filter_3))
                                             )
                                     )]["ID"].tolist()

        k09_wildfl_filter = ["Wild animal, game (Includes birds; not deer or elk)", "Deer or elk, wapiti"]
        k09_raw_df = raw_df[(raw_df["Crash Level Event 1 Code"].isin(k09_wildfl_filter)) |
                            (raw_df["Crash Level Event 2 Code"].isin(k09_wildfl_filter)) |
                            (raw_df["Crash Level Event 3 Code"].isin(k09_wildfl_filter)) |
                            (raw_df["Vehicle Event 1 Code"].isin(k09_wildfl_filter)) |
                            (raw_df["Vehicle Event 2 Code"].isin(k09_wildfl_filter)) |
                            (raw_df["Vehicle Event 3 Code"].isin(k09_wildfl_filter))]
        k09_wildfl_filter_2 = k09_raw_df["ID"].tolist()

        k10_disdfl_filter_1 = ["Other (phantom) non-contact vehicle (On PAR or witness statement)",
                               "Cell phone (on PAR or report submitted by driver using phone)",
                               "Cell phone use witnessed by other participant", "Texting ",
                               "Distracted by navigation system or GPS device", "Distracted by other electronic device",
                               "Passenger interfered with driver","Animal or insect in vehicle interfered with driver"]
        k10_disdfl_filter_2 = ["Passenger interfering with driver","Driver’s attention distracted","Inattention","Property damage only crash (PDO)"]
        k10_disdfl_1 = raw_df[raw_df['Participant Action'] == k10_disdfl_filter_2[0]]["ID"].tolist()

        k10_disdfl_2 = raw_df[(raw_df['Participant Type Code'] == 'Driver') &
                              ((raw_df['Participant Action'] == k10_disdfl_filter_2[1]) |
                               (raw_df['Participant Cause 1 Code'] == k10_disdfl_filter_2[2]) |
                               (raw_df['Participant Cause 2 Code'] == k10_disdfl_filter_2[2]) |
                               (raw_df['Participant Cause 3 Code'] == k10_disdfl_filter_2[2]) |
                               (raw_df['Participant Error 1 Code'] == k10_disdfl_filter_2[2]) |
                               (raw_df['Participant Error 2 Code'] == k10_disdfl_filter_2[2]) |
                               (raw_df['Participant Error 3 Code'] == k10_disdfl_filter_2[2]) |
                               (raw_df['Participant Event 1 Code'].isin(k10_disdfl_filter_1)) |
                               (raw_df['Participant Event 2 Code'].isin(k10_disdfl_filter_1)) |
                               (raw_df['Participant Event 3 Code'].isin(k10_disdfl_filter_1))
                               )]["ID"].tolist()

        k10_disdfl_3 = raw_df[(raw_df['Crash Year'] >= 2016) &
                              ((raw_df['Crash Severity'] == k10_disdfl_filter_2[3]) |
                               (raw_df['Crash Level Cause 1 Code'] == k10_disdfl_filter_2[2]) |
                               (raw_df['Crash Level Cause 2 Code'] == k10_disdfl_filter_2[2]) |
                               (raw_df['Crash Level Cause 3 Code'] == k10_disdfl_filter_2[2]) |
                               (raw_df['Crash Level Event 1 Code'].isin(k10_disdfl_filter_1)) |
                               (raw_df['Crash Level Event 2 Code'].isin(k10_disdfl_filter_1)) |
                               (raw_df['Crash Level Event 3 Code'].isin(k10_disdfl_filter_1))
                               )]["ID"].tolist()

        k10_disdfl_filter_3 = k10_disdfl_1 + k10_disdfl_2 + k10_disdfl_3

        k13_pedfl_filter = ["Pedestrian indirectly involved (Not struck)", "Hitchhiker (soliciting a ride)"]
        k13_raw_df = raw_df[(raw_df["Crash Level Event 1 Code"].isin(k13_pedfl_filter)) |
                            (raw_df["Crash Level Event 2 Code"].isin(k13_pedfl_filter)) |
                            (raw_df["Crash Level Event 3 Code"].isin(k13_pedfl_filter)) |
                            (raw_df["Vehicle Event 1 Code"].isin(k13_pedfl_filter)) |
                            (raw_df["Vehicle Event 2 Code"].isin(k13_pedfl_filter)) |
                            (raw_df["Vehicle Event 3 Code"].isin(k13_pedfl_filter))]
        k13_pedfl_filter_2 = k13_raw_df["ID"].tolist()

        k14_pcycfl_filter = ["Pedal-cyclist indirectly involved (Not struck)"]
        k14_raw_df = raw_df[(raw_df["Crash Level Event 1 Code"] == k14_pcycfl_filter[0]) |
                            (raw_df["Crash Level Event 2 Code"] == k14_pcycfl_filter[0]) |
                            (raw_df["Crash Level Event 3 Code"] == k14_pcycfl_filter[0]) |
                            (raw_df["Vehicle Event 1 Code"] == k14_pcycfl_filter[0]) |
                            (raw_df["Vehicle Event 2 Code"] == k14_pcycfl_filter[0]) |
                            (raw_df["Vehicle Event 3 Code"] == k14_pcycfl_filter[0])]
        k14_pcycfl_filter_2 = k14_raw_df["ID"].tolist()

        k62_trnfl_filter_1 = ["Railway train","Train struck vehicle","Vehicle struck train"]
        k62_trnfl_filter_2 = raw_df[(raw_df["Crash Type"] == k62_trnfl_filter_1[0]) |
                                    (raw_df["Crash Level Event 1 Code"] == k62_trnfl_filter_1[1]) |
                                    (raw_df["Crash Level Event 2 Code"] == k62_trnfl_filter_1[1]) |
                                    (raw_df["Crash Level Event 3 Code"] == k62_trnfl_filter_1[1]) |
                                    (raw_df["Vehicle Event 1 Code"] == k62_trnfl_filter_1[1]) |
                                    (raw_df["Vehicle Event 2 Code"] == k62_trnfl_filter_1[1]) |
                                    (raw_df["Vehicle Event 3 Code"] == k62_trnfl_filter_1[1]) |
                                    (raw_df["Crash Level Event 1 Code"] == k62_trnfl_filter_1[2]) |
                                    (raw_df["Crash Level Event 2 Code"] == k62_trnfl_filter_1[2]) |
                                    (raw_df["Crash Level Event 3 Code"] == k62_trnfl_filter_1[2]) |
                                    (raw_df["Vehicle Event 1 Code"] == k62_trnfl_filter_1[2]) |
                                    (raw_df["Vehicle Event 2 Code"] == k62_trnfl_filter_1[2]) |
                                    (raw_df["Vehicle Event 3 Code"] == k62_trnfl_filter_1[2])]["ID"].tolist()

        k63_protfl_filter_1 = ["No safety equipment used","Seat belt or harness used improperly","Child restraint used improperly",
                               "Helmet used improperly"]
        k63_protfl_filter_2 = raw_df[raw_df["Participant Safety Equipment Use Code"].isin(k63_protfl_filter_1)]["ID"].tolist()

        for i in range(len(raw_df["ID"])):
            if raw_df["Record Type"][i] == 1:
                if raw_df["Total Fatality Count"][i] > 0:
                    k01_sevr.append("Fatal")
                    k02_sevrcl.append("Fatal/Severe Injury")
                    k03_sevrfl.append(1)
                    k04_epdo.append(100)
                elif raw_df["Total Suspected Serious Injury (A) Count"][i] > 0:
                    k01_sevr.append("Severe Injury")
                    k02_sevrcl.append("Fatal/Severe Injury")
                    k03_sevrfl.append(1)
                    k04_epdo.append(100)
                elif raw_df["Total Suspected Minor Injury (B) Count"][i] > 0:
                    k01_sevr.append("Moderate Injury")
                    k02_sevrcl.append("Other Injury")
                    k03_sevrfl.append(0)
                    k04_epdo.append(10)
                elif raw_df["Total Possible Injury (C) Count"][i] > 0:
                    k01_sevr.append("Minor Injury")
                    k02_sevrcl.append("Other Injury")
                    k03_sevrfl.append(0)
                    k04_epdo.append(10)
                else:
                    k01_sevr.append("PDO")
                    k02_sevrcl.append("PDO")
                    k03_sevrfl.append(0)
                    k04_epdo.append(1)
                if raw_df["Alcohol-Involved Flag"][i] == "Yes" or raw_df["Drugs Involved Flag"][i] == "Yes" or raw_df["Crash Marijuana Involved Flag"][i] == "Yes":
                    k06_imprfl.append(1)
                else:
                    k06_imprfl.append(0)

                if raw_df["Crash Level Cause 1 Code"][i] == "*Aggressive Driving (Per PAR)" or raw_df["Crash Level Cause 2 Code"][i] == "*Aggressive Driving (Per PAR)" or raw_df["Crash Level Cause 2 Code"][i] == "*Aggressive Driving (Per PAR)":
                    k07_aggfl.append(1)
                else:
                    k07_aggfl.append(0)

                if len(k08_rdepfl_filter_4) != 0:
                    if raw_df["ID"][i] in k08_rdepfl_filter_4:
                        k08_rdepfl.append(1)
                    else:
                        k08_rdepfl.append(0)
                else:
                    k08_rdepfl.append(0)

                if len(k09_wildfl_filter_2) != 0:
                    if raw_df["ID"][i] in k09_wildfl_filter_2:
                        k09_wildfl.append(1)
                    else:
                        k09_wildfl.append(0)
                else:
                    k09_wildfl.append(0)

                if len(k10_disdfl_filter_3) != 0:
                    if raw_df["ID"][i] in k10_disdfl_filter_3:
                        k10_disdfl.append(1)
                    else:
                        k10_disdfl.append(0)
                else:
                    k10_disdfl.append(0)

                if raw_df["Road Character"][i] == "Street/road or highway intersection" or raw_df["Intersection Related Flag"][i] == "Yes":
                    k11_intxfl.append(1)
                else:
                    k11_intxfl.append(0)

                # k12_motrfl_filter = ["Motorcycle, dirt bike"]
                # if raw_df["Vehicle Type Code"][i] in k12_motrfl_filter:
                #     k12_motrfl.append(1)
                # else:
                #     k12_motrfl.append(0)

                if len(k13_pedfl_filter_2) != 0:
                    if raw_df["Total Pedestrian Count"][i] == 0 and raw_df["ID"][i] in k13_pedfl_filter_2:
                        k13_pedfl.append(1)
                    else:
                        k13_pedfl.append(0)
                else:
                    k13_pedfl.append(0)

                if len(k14_pcycfl_filter_2) != 0:

                    if raw_df["Total Pedalcyclist Count"][i] == 0 and raw_df["ID"][i] in k14_pcycfl_filter_2:
                        k14_pcycfl.append(1)
                    else:
                        k14_pcycfl.append(0)
                else:
                    k14_pcycfl.append(0)

                if len(k62_trnfl_filter_2) != 0:
                    if raw_df["ID"][i] in k62_trnfl_filter_2:
                        k62_trnfl.append(1)
                    else:
                        k62_trnfl.append(0)
                else:
                    k62_trnfl.append(0)

                if len(k63_protfl_filter_2) != 0:
                    if raw_df["ID"][i] in k63_protfl_filter_2:
                        k63_protfl.append(1)
                    else:
                        k63_protfl.append(0)
                else:
                    k63_protfl.append(0)

                if raw_df["Total Pedestrian Count"][i] > 0:
                    k65_peddi.append(1)
                else:
                    k65_peddi.append(0)

                if raw_df["Total Pedalcyclist Count"][i] > 0:
                    k66_pcycdi.append(1)
                else:
                    k66_pcycdi.append(0)

            # elif raw_df["Record Type"][i] == 2:
            #     k12_motrfl_filter = ["Motorcycle, dirt bike"]
            #     if raw_df["Vehicle Type Code"][i] in k12_motrfl_filter:
            #         k12_motrfl.append(1)
            #     else:
            #         k12_motrfl.append(0)

            else:
                k01_sevr.append(np.nan)
                k02_sevrcl.append(np.nan)
                k03_sevrfl.append(np.nan)
                k04_epdo.append(np.nan)
                k06_imprfl.append(np.nan)
                k07_aggfl.append(np.nan)
                k08_rdepfl.append(np.nan)
                k09_wildfl.append(np.nan)
                k10_disdfl.append(np.nan)
                k11_intxfl.append(np.nan)
                # k12_motrfl.append(np.nan)
                k13_pedfl.append(np.nan)
                k14_pcycfl.append(np.nan)
                k62_trnfl.append(np.nan)
                k63_protfl.append(np.nan)
                k66_pcycdi.append(np.nan)
                k65_peddi.append(np.nan)

        motrorcycle_fl = raw_df[raw_df["Record Type"] == 2].copy()
        if len(motrorcycle_fl) > 0:
            motrorcycle_fl.loc[motrorcycle_fl["Vehicle Type Code"] == "Motorcycle, dirt bike", 'k12_motrfl'] = 1
            motrorcycle_fl.loc[motrorcycle_fl["Vehicle Type Code"] != "Motorcycle, dirt bike", 'k12_motrfl'] = 0

        # Youngest & Older Driver Age
        k15_ydage = raw_df[(raw_df['Participant Type Code'] == "Driver") & (raw_df['Age'] != 0)]
        if len(k15_ydage) > 0:

            k15_ydage = k15_ydage.loc[k15_ydage.groupby('ID')['Age'].idxmin().reset_index(drop=True)]
            k15_ydage = k15_ydage.rename(columns={'Age': 'k15_ydage'})

        k16_odage = raw_df[(raw_df['Participant Type Code'] == "Driver") & (raw_df['Age'] != 0)]
        if len(k16_odage) > 0:
            k16_odage = k16_odage.loc[k16_odage.groupby('ID')['Age'].idxmax().reset_index(drop=True)]
            k16_odage = k16_odage.rename(columns={'Age': 'k16_odage'})

        # """
        #     Youngest & Older Driver at Error Age
        # """
        k17_ydeage = raw_df[(raw_df['Participant Type Code'] == "Driver") & (raw_df['Age'] != 0) & (raw_df['Participant Error 1 Code'] != "No error") &
                            ((~raw_df['Participant Error 1 Code'].isnull())|(~raw_df['Participant Error 2 Code'].isnull())|(~raw_df['Participant Error 3 Code'].isnull()))]
        if len(k17_ydeage) > 0:
            k17_ydeage = k17_ydeage.loc[k17_ydeage.groupby('ID')['Age'].idxmin().reset_index(drop=True)]
            k17_ydeage = k17_ydeage.rename(columns={'Age': 'k17_ydeage'})

            # """
            # Young driver error flag
            # """

            k17_ydeage.loc[(k17_ydeage["k17_ydeage"] >= 15) & (k17_ydeage["k17_ydeage"] <= 20), 'k20_yderfl'] = 1
            k17_ydeage.loc[k17_ydeage["k17_ydeage"] > 20, 'k20_yderfl'] = 0

        k18_odeage = raw_df[(raw_df['Participant Type Code'] == "Driver") & (raw_df['Age'] != 0) & (raw_df['Participant Error 1 Code'] != "No error") &
                            ((~raw_df['Participant Error 1 Code'].isnull())|(~raw_df['Participant Error 2 Code'].isnull())|(~raw_df['Participant Error 3 Code'].isnull()))]
        if len(k18_odeage) > 0:
            k18_odeage = k18_odeage.loc[k18_odeage.groupby('ID')['Age'].idxmax().reset_index(drop=True)]
            k18_odeage = k18_odeage.rename(columns={'Age': 'k18_odeage'})

            # """
            # Old driver error flag
            # """
            k18_odeage.loc[(k18_odeage["k18_odeage"] >= 65), 'k22_odefl'] = 1
            k18_odeage.loc[k18_odeage["k18_odeage"] < 65, 'k22_odefl'] = 0

        # """
        # Number of drivers at error
        # """
        k19_nder = raw_df[(raw_df['Participant Type Code'] == "Driver") & (raw_df['Participant Error 1 Code'] != "No error") &
                          ((~raw_df['Participant Error 1 Code'].isnull())|(~raw_df['Participant Error 2 Code'].isnull())|(~raw_df['Participant Error 3 Code'].isnull()))]
        if len(k19_nder) > 0:
            k19_nder = k19_nder.groupby('ID').agg({'Participant Type Code':['count']}).copy()
            k19_nder.columns = k19_nder.columns.map('|'.join).str.strip('|')
            k19_nder.reset_index(inplace=True)
            k19_nder = k19_nder.rename(columns={'Participant Type Code|count': 'k19_nder'})
            k19_nder["Participant Vehicle ID"] = "General Crash Information"

        # """
        # young driver & Old driver flag
        # """
        k21_ydfl = raw_df[(raw_df['Participant Type Code'] == "Driver") & (raw_df['Age'] != 0)].copy()
        if len(k21_ydfl) > 0:
            k21_ydfl.loc[(k21_ydfl['Age'] >= 15) & (k21_ydfl["Age"] <= 20), 'k21_ydfl'] = 1
            k21_ydfl.loc[(k21_ydfl['Age'] < 15) | (k21_ydfl["Age"] > 20), 'k21_ydfl'] = 0

        k23_odfl = raw_df[(raw_df['Participant Type Code'] == "Driver") & (raw_df['Age'] != 0)].copy()
        if len(k23_odfl) > 0:
            k23_odfl.loc[(k23_odfl['Age'] >= 65),'k23_odfl'] = 1
            k23_odfl.loc[(k23_odfl['Age'] < 65), 'k23_odfl'] = 0

        # """
        # young Ped & Old Ped age
        # """
        k24_ypage_filter = ["Pedestrian", "Pedestrian using a pedestrian conveyance (wheelchair, skates, etc.)", "Pedestrian towing an object, other participant"]
        k24_ypage = raw_df[(raw_df['Participant Type Code'].isin(k24_ypage_filter)) & (raw_df['Age'] != 0)]
        if len(k24_ypage) > 0:
            k24_ypage = k24_ypage.loc[k24_ypage.groupby('ID')['Age'].idxmin().reset_index(drop=True)]
            k24_ypage = k24_ypage.rename(columns={'Age': 'k24_ypage'})

        k25_opage_filter = ["Pedestrian", "Pedestrian using a pedestrian conveyance (wheelchair, skates, etc.)", "Pedestrian towing an object, other participant"]
        k25_opage = raw_df[(raw_df['Participant Type Code'].isin(k25_opage_filter)) & (raw_df['Age'] != 0)]
        if len(k25_opage) > 0:
            k25_opage = k25_opage.loc[k25_opage.groupby('ID')['Age'].idxmax().reset_index(drop=True)]
            k25_opage = k25_opage.rename(columns={'Age': 'k25_opage'})

        # """
        # Male or Female or non binary or unknown driver involved
        # """
        driver_sex = raw_df[raw_df['Participant Type Code'] == "Driver"].copy()
        if len(driver_sex) > 0:
            driver_sex.loc[driver_sex["Sex"] == "Male", 'k26_mldr'] = 1
            driver_sex.loc[driver_sex["Sex"] != "Male", 'k26_mldr'] = 0
            driver_sex.loc[driver_sex["Sex"] == "Female", 'k27_fmdr'] = 1
            driver_sex.loc[driver_sex["Sex"] != "Female", 'k27_fmdr'] = 0
            driver_sex.loc[driver_sex["Sex"] == "Non-Binary Gender", 'k28_nbdr'] = 1
            driver_sex.loc[driver_sex["Sex"] != "Non-Binary Gender", 'k28_nbdr'] = 0
            driver_sex.loc[driver_sex["Sex"] == "Unknown", 'k33_ukdr'] = 1
            driver_sex.loc[driver_sex["Sex"] != "Unknown", 'k33_ukdr'] = 0

        # """
        # Male or Female or non binary or unknown driver at error
        # """
        driver = raw_df[(raw_df['Participant Type Code'] == "Driver")].copy()
        if len(driver) > 0:
            driver.loc[(driver["Sex"] == "Male") & (driver['Participant Error 1 Code'] != "No error") & ((~driver['Participant Error 1 Code'].isnull())|(~driver['Participant Error 2 Code'].isnull())|(~driver['Participant Error 3 Code'].isnull())),'k29_mlder'] = 1
            driver.loc[(driver["Sex"] == "Male") & (driver['Participant Error 1 Code'] == "No error"),'k29_mlder'] = 0
            driver.loc[(driver["Sex"] == "Female") & (driver['Participant Error 1 Code'] != "No error") & ((~driver['Participant Error 1 Code'].isnull())|(~driver['Participant Error 2 Code'].isnull())|(~driver['Participant Error 3 Code'].isnull())),'k30_fmder'] = 1
            driver.loc[(driver["Sex"] == "Female") & (driver['Participant Error 1 Code'] == "No error"),'k30_fmder'] = 0
            driver.loc[(driver["Sex"] == "Non-Binary Gender") & (driver['Participant Error 1 Code'] != "No error") & ((~driver['Participant Error 1 Code'].isnull())|(~driver['Participant Error 2 Code'].isnull())|(~driver['Participant Error 3 Code'].isnull())),'k31_nbder'] = 1
            driver.loc[(driver["Sex"] == "Non-Binary Gender") & (driver['Participant Error 1 Code'] == "No error"),'k31_nbder'] = 0
            driver.loc[(driver["Sex"] == "Unknown") & (driver['Participant Error 1 Code'] != "No error") & ((~driver['Participant Error 1 Code'].isnull())|(~driver['Participant Error 2 Code'].isnull())|(~driver['Participant Error 3 Code'].isnull())),'k32_ukder'] = 1
            driver.loc[(driver["Sex"] == "Unknown") & (driver['Participant Error 1 Code'] == "No error"),'k32_ukder'] = 0
        # """
        #     Youngest & Older Bicyclist Age
        #     """
        k34_ybage_filter = ["Pedalcyclist", "Pedalcyclist towing an object, other participant"]
        k34_ybage = raw_df[(raw_df['Participant Type Code'].isin(k34_ybage_filter)) & (raw_df['Age'] != 0)]
        if len(k34_ybage) > 0:
            k34_ybage = k34_ybage.loc[k34_ybage.groupby('ID')['Age'].idxmin().reset_index(drop=True)]
            k34_ybage = k34_ybage.rename(columns={'Age': 'k34_ybage'})

        k35_obage_filter = ["Pedalcyclist", "Pedalcyclist towing an object, other participant"]
        k35_obage = raw_df[(raw_df['Participant Type Code'].isin(k35_obage_filter)) & (raw_df['Age'] != 0)]
        if len(k35_obage) > 0:
            k35_obage = k35_obage.loc[k35_obage.groupby('ID')['Age'].idxmax().reset_index(drop=True)]
            k35_obage = k35_obage.rename(columns={'Age': 'k35_obage'})

        # """
        # Male or Female or non binary or unknown bicyclist
        # """
        bicyclist_sex_filter = ["Pedalcyclist", "Pedalcyclist towing an object, other participant"]
        bicyclist_sex = raw_df[raw_df['Participant Type Code'].isin(bicyclist_sex_filter)].copy()
        if len(bicyclist_sex) > 0:
            bicyclist_sex.loc[bicyclist_sex["Sex"] == "Female", 'k36_fmbage'] = 1
            bicyclist_sex.loc[bicyclist_sex["Sex"] != "Female", 'k36_fmbage'] = 0
            bicyclist_sex.loc[bicyclist_sex["Sex"] == "Male", 'k37_mlbage'] = 1
            bicyclist_sex.loc[bicyclist_sex["Sex"] != "Male", 'k37_mlbage'] = 0
            bicyclist_sex.loc[bicyclist_sex["Sex"] == "Non-Binary Gender", 'k38_nbbage'] = 1
            bicyclist_sex.loc[bicyclist_sex["Sex"] != "Non-Binary Gender", 'k38_nbbage'] = 0
            bicyclist_sex.loc[bicyclist_sex["Sex"] == "Unknown", 'k39_ukbage'] = 1
            bicyclist_sex.loc[bicyclist_sex["Sex"] != "Unknown", 'k39_ukbage'] = 0

        # """
        # Male or Female or non binary or unknown Ped Involved
        # """
        ped_sex_filter = ["Pedestrian", "Pedestrian using a pedestrian conveyance (wheelchair, skates, etc.)", "Pedestrian towing an object, other participant"]
        ped_sex = raw_df[raw_df['Participant Type Code'].isin(ped_sex_filter)].copy()
        if len(ped_sex) > 0:
            ped_sex.loc[ped_sex["Sex"] == "Female", 'k40_fmped'] = 1
            ped_sex.loc[ped_sex["Sex"] != "Female", 'k40_fmped'] = 0
            ped_sex.loc[ped_sex["Sex"] == "Male", 'k41_mlped'] = 1
            ped_sex.loc[ped_sex["Sex"] != "Male", 'k41_mlped'] = 0
            ped_sex.loc[ped_sex["Sex"] == "Non-Binary Gender", 'k42_nbped'] = 1
            ped_sex.loc[ped_sex["Sex"] != "Non-Binary Gender", 'k42_nbped'] = 0
            ped_sex.loc[ped_sex["Sex"] == "Unknown", 'k43_ukped'] = 1
            ped_sex.loc[ped_sex["Sex"] != "Unknown", 'k43_ukped'] = 0

        # """ 
        # Vehicle Travel Direction
        # """
        k44_dirvh = raw_df[(~raw_df["Vehicle Travel Direction From"].isnull()) & (~raw_df["Vehicle Travel Direction To"].isnull())].copy()
        if len(k44_dirvh) > 0:
            k44_dirvh["k44_dirvh"] = "From " + k44_dirvh["Vehicle Travel Direction From"].astype(str) + " To " + k44_dirvh["Vehicle Travel Direction To"].astype(str)

        # """
        # Driver Residence    
        # """
        driver_res = raw_df[(raw_df['Participant Type Code'] == "Driver")].copy()
        if len(driver_res) > 0:
            driver_res.loc[driver_res["Driver Residence Status"] == "OR Res. <25 mi of home", "k45_lcdfl"] = 1
            driver_res.loc[driver_res["Driver Residence Status"] != "OR Res. <25 mi of home", "k45_lcdfl"] = 0

            driver_res.loc[driver_res["Driver Residence Status"] == "OR Res. >25 mi of home", "k46_nlcdfl"] = 1
            driver_res.loc[driver_res["Driver Residence Status"] != "OR Res. >25 mi of home", "k46_nlcdfl"] = 0

        # """
        # Total Vehicles Involved,  Vehicle classification
        # """
        k47_numvh = raw_df[raw_df["Record Type"] == 2].copy()
        k55_vhclss = pd.DataFrame()
        if len(k47_numvh) > 0:
            k47_numvh = k47_numvh.groupby("ID").agg({"Record Type": 'count'}).copy()
            k47_numvh.reset_index(inplace=True)
            k47_numvh = k47_numvh.rename(columns={'Record Type': 'k47_numvh'})
            k47_numvh["Participant Vehicle ID"] = "General Crash Information"

            k55_vhclss = k47_numvh[k47_numvh["k47_numvh"] != 0].copy()
            if len(k55_vhclss) > 0:
                k55_vhclss.loc[k55_vhclss["k47_numvh"] == 0, 'k55_vhclss'] = "no vehicle"
                k55_vhclss.loc[k55_vhclss["k47_numvh"] == 1 , 'k55_vhclss'] = "single vehicle"
                k55_vhclss.loc[k55_vhclss["k47_numvh"] > 1, 'k55_vhclss'] = "multi-vehicle"

        # """
        # Date, month, time, hour
        # """
        k48_crshdt = raw_df[raw_df["Record Type"] == 1].copy()
        if len(k48_crshdt) > 0:
            k48_crshdt["k48_crshdt"] = k48_crshdt["Crash Year"].astype(int).astype(str)+"-"+k48_crshdt["Crash Month"].astype(int).astype(str)+"-"+k48_crshdt["Crash Day"].astype(int).astype(str)

        k49_crshmo = raw_df[raw_df["Record Type"] == 1].copy()
        if len(k49_crshmo) > 0:
            k49_crshmo.loc[k49_crshmo["Crash Month"] == 1,  'k49_crshmo'] = "January"
            k49_crshmo.loc[k49_crshmo["Crash Month"] == 2, 'k49_crshmo'] = "February"
            k49_crshmo.loc[k49_crshmo["Crash Month"] == 3, 'k49_crshmo'] = "March"
            k49_crshmo.loc[k49_crshmo["Crash Month"] == 4, 'k49_crshmo'] = "April"
            k49_crshmo.loc[k49_crshmo["Crash Month"] == 5, 'k49_crshmo'] = "May"
            k49_crshmo.loc[k49_crshmo["Crash Month"] == 6, 'k49_crshmo'] = "June"
            k49_crshmo.loc[k49_crshmo["Crash Month"] == 7, 'k49_crshmo'] = "July"
            k49_crshmo.loc[k49_crshmo["Crash Month"] == 8, 'k49_crshmo'] = "August"
            k49_crshmo.loc[k49_crshmo["Crash Month"] == 9, 'k49_crshmo'] = "September"
            k49_crshmo.loc[k49_crshmo["Crash Month"] == 10, 'k49_crshmo'] = "October"
            k49_crshmo.loc[k49_crshmo["Crash Month"] == 11, 'k49_crshmo'] = "November"
            k49_crshmo.loc[k49_crshmo["Crash Month"] == 12, 'k49_crshmo'] = "December"

        k50_crshtm = raw_df[(raw_df["Record Type"] == 1) & (~raw_df["Crash Hour"].isnull())].copy()
        if len(k50_crshtm) > 0:
            k50_crshtm["k50_crshtm"] = k50_crshtm["Crash Hour"].astype(int).astype(str) + ":00"

        # """ 
        # Collision
        # """
        k54_coltyp = raw_df[raw_df["Record Type"] == 1].copy()
        if len(k54_coltyp) > 0:
            k54_coltyp["k54_coltyp"] = k54_coltyp["Collision Type"]
            k54_coltyp.loc[k54_coltyp["Crash Type"] == "Pedalcyclist", 'k54_coltyp'] = "Pedalcyclist"
        # """
        # Sign /Signal Violation
        # """
        k59_siviol_filter =["Failed to obey mandatory traffic turn signal, sign or lane markings", "Disregarded traffic signal",
                            "Disregarded stop sign or flashing red","Disregarded warning sign, flares or flashing amber",
                            "Disregarded Rail Road signal, Rail Road sign, or Rail Road flagman"]
        k59_siviol = raw_df[(raw_df['Participant Type Code'] == "Driver")].copy()
        if len(k59_siviol) > 0:
            k59_siviol.loc[(k59_siviol["Participant Error 1 Code"].isin(k59_siviol_filter)) | (k59_siviol["Participant Error 2 Code"].isin(k59_siviol_filter)) | (k59_siviol["Participant Error 3 Code"].isin(k59_siviol_filter)), 'k59_siviol'] = 1
            k59_siviol.loc[(~k59_siviol["Participant Error 1 Code"].isin(k59_siviol_filter)) & (~k59_siviol["Participant Error 2 Code"].isin(k59_siviol_filter)) & (~k59_siviol["Participant Error 3 Code"].isin(k59_siviol_filter)), 'k59_siviol'] = 0

        # """
        # Crash City / Place
        # """
        k61_crshpl = raw_df[raw_df["Record Type"] == 1].reset_index().copy()
        if len(k61_crshpl) > 0:
            k61_crshpl["k61_crshpl"] = "N/A"
            for k61_idx in range(len(k61_crshpl["Record Type"])):
                if str(k61_crshpl["City Section ID"][k61_idx]) != "nan":
                    k61_crshpl.loc[k61_idx,'k61_crshpl'] = k61_crshpl["City Section ID"][k61_idx]
                elif str(k61_crshpl["Urban Area Code"][k61_idx]) != "nan":
                    k61_crshpl.loc[k61_idx,'k61_crshpl'] = k61_crshpl["Urban Area Code"][k61_idx]
                elif str(k61_crshpl["County Code"][k61_idx]) != "nan":
                    k61_crshpl.loc[k61_idx,'k61_crshpl'] = k61_crshpl["County Code"][k61_idx]

        num_veh_dir = raw_df.groupby(["ID",'Vehicle Travel Direction To']).agg({"Vehicle Travel Direction To": ['count']}).copy()
        num_veh_dir.columns = num_veh_dir.columns.map('|'.join).str.strip('|')
        num_veh_dir.reset_index(inplace=True)
        num_veh_dir = num_veh_dir.pivot(index="ID", columns="Vehicle Travel Direction To", values="Vehicle Travel Direction To|count").reset_index()
        num_veh_dir["Participant Vehicle ID"] = "General Crash Information"
        num_veh_dir.rename(columns={'north': 'k67_nvh', 'east': 'k68_evh', 'south': 'k69_svh', 'west': 'k70_wvh',
                                     'northeast': 'k71_nevh', 'southeast': 'k72_sevh', 'southwest': 'k73_swvh',
                                     'northwest': 'k74_nwvh', 'unknown': 'k75_ukvh'}, inplace=True)
        num_veh_dir_cols = list(num_veh_dir.columns.values)
        num_veh_dir_cols.sort()
        num_veh_dir = num_veh_dir.reindex(columns=num_veh_dir_cols)

        raw_df["k01_sevr"] = k01_sevr
        raw_df["k02_sevrcl"] = k02_sevrcl
        raw_df["k03_sevrfl"] = k03_sevrfl
        raw_df["k04_epdo"] = k04_epdo
        raw_df["k06_imprfl"] = k06_imprfl
        raw_df["k07_aggfl"] = k07_aggfl
        raw_df["k08_rdepfl"] = k08_rdepfl
        raw_df["k09_wildfl"] = k09_wildfl
        raw_df["k10_disdfl"] = k10_disdfl
        raw_df["k11_intxfl"] = k11_intxfl

        if len(motrorcycle_fl) > 0:
            raw_df = raw_df.merge(motrorcycle_fl[['Participant Vehicle ID','ID','k12_motrfl']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k12_motrfl"] = [np.nan] * len(raw_df["ID"])

        raw_df["k13_pedfl"] = k13_pedfl
        raw_df["k14_pcycfl"] = k14_pcycfl

        if len(k15_ydage) > 0:
            raw_df = raw_df.merge(k15_ydage[['Participant Vehicle ID','ID','k15_ydage']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k15_ydage"] = [np.nan] * len(raw_df["ID"])

        if len(k16_odage) > 0:
            raw_df = raw_df.merge(k16_odage[['Participant Vehicle ID','ID','k16_odage']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k16_odage"] = [np.nan] * len(raw_df["ID"])

        if len(k17_ydeage) > 0:
            raw_df = raw_df.merge(k17_ydeage[['Participant Vehicle ID','ID','k17_ydeage']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k17_ydeage"] = [np.nan] * len(raw_df["ID"])

        if len(k18_odeage) > 0:
            raw_df = raw_df.merge(k18_odeage[['Participant Vehicle ID','ID','k18_odeage']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k18_odeage"] = [np.nan] * len(raw_df["ID"])
        if len(k19_nder) > 0:
            raw_df = raw_df.merge(k19_nder[['Participant Vehicle ID','ID','k19_nder']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k19_nder"] = [np.nan] * len(raw_df["ID"])

        if len(k17_ydeage) > 0:
            raw_df = raw_df.merge(k17_ydeage[['Participant Vehicle ID','ID','k20_yderfl']], on=['ID','Participant Vehicle ID'], how= 'left') #k20
        else:
            raw_df["k20_yderfl"] = [np.nan] * len(raw_df["ID"])

        if len(k21_ydfl) > 0:
            raw_df = raw_df.merge(k21_ydfl.reset_index()[['Participant Vehicle ID','ID','k21_ydfl']], on=['ID','Participant Vehicle ID'], how= 'left')

        else:
            raw_df["k21_ydfl"] = [np.nan] * len(raw_df["ID"])

        if len(k18_odeage) > 0:
            raw_df = raw_df.merge(k18_odeage[['Participant Vehicle ID','ID','k22_odefl']], on=['ID','Participant Vehicle ID'], how= 'left') #k22
        else:
            raw_df["k22_odefl"] = [np.nan] * len(raw_df["ID"])

        if len(k23_odfl) > 0:
            raw_df = raw_df.merge(k23_odfl[['Participant Vehicle ID','ID','k23_odfl']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k23_odfl"] = [np.nan] * len(raw_df["ID"])

        if len(k24_ypage) > 0:
            raw_df = raw_df.merge(k24_ypage[['Participant Vehicle ID','ID','k24_ypage']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k24_ypage"] = [np.nan] * len(raw_df["ID"])
        if len(k25_opage) > 0:
            raw_df = raw_df.merge(k25_opage[['Participant Vehicle ID','ID','k25_opage']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k25_opage"] = [np.nan] * len(raw_df["ID"])


        if len(driver_sex) > 0:
            raw_df = raw_df.merge(driver_sex[['Participant Vehicle ID','ID','k26_mldr']], on=['ID','Participant Vehicle ID'], how= 'left')
            raw_df = raw_df.merge(driver_sex[['Participant Vehicle ID','ID','k27_fmdr']], on=['ID','Participant Vehicle ID'], how= 'left')
            raw_df = raw_df.merge(driver_sex[['Participant Vehicle ID','ID','k28_nbdr']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k26_mldr"] = [np.nan] * len(raw_df["ID"])
            raw_df["k27_fmdr"] = [np.nan] * len(raw_df["ID"])
            raw_df["k28_nbdr"] = [np.nan] * len(raw_df["ID"])

        if len(driver) > 0:
            raw_df = raw_df.merge(driver[['Participant Vehicle ID','ID','k29_mlder']], on=['ID','Participant Vehicle ID'], how= 'left')
            raw_df = raw_df.merge(driver[['Participant Vehicle ID','ID','k30_fmder']], on=['ID','Participant Vehicle ID'], how= 'left')
            raw_df = raw_df.merge(driver[['Participant Vehicle ID','ID','k31_nbder']], on=['ID','Participant Vehicle ID'], how= 'left')
            raw_df = raw_df.merge(driver[['Participant Vehicle ID','ID','k32_ukder']], on=['ID','Participant Vehicle ID'], how= 'left')
            raw_df = raw_df.merge(driver_sex[['Participant Vehicle ID','ID','k33_ukdr']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k29_mlder"] = [np.nan] * len(raw_df["ID"])
            raw_df["k30_fmder"] = [np.nan] * len(raw_df["ID"])
            raw_df["k31_nbder"] = [np.nan] * len(raw_df["ID"])
            raw_df["k32_ukder"] = [np.nan] * len(raw_df["ID"])
            raw_df["k33_ukdr"] = [np.nan] * len(raw_df["ID"])

        if len(k34_ybage) > 0:
            raw_df = raw_df.merge(k34_ybage[['Participant Vehicle ID','ID','k34_ybage']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k34_ybage"] = [np.nan] * len(raw_df["ID"])
        if len(k35_obage) > 0:
            raw_df = raw_df.merge(k35_obage[['Participant Vehicle ID','ID','k35_obage']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k35_obage"] = [np.nan] * len(raw_df["ID"])
        if len(bicyclist_sex) > 0:
            raw_df = raw_df.merge(bicyclist_sex[['Participant Vehicle ID','ID','k36_fmbage']], on=['ID','Participant Vehicle ID'], how= 'left')
            raw_df = raw_df.merge(bicyclist_sex[['Participant Vehicle ID','ID','k37_mlbage']], on=['ID','Participant Vehicle ID'], how= 'left')
            raw_df = raw_df.merge(bicyclist_sex[['Participant Vehicle ID', 'ID', 'k38_nbbage']], on=['ID', 'Participant Vehicle ID'], how='left')
            raw_df = raw_df.merge(bicyclist_sex[['Participant Vehicle ID','ID','k39_ukbage']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k36_fmbage"] = [np.nan] * len(raw_df["ID"])
            raw_df["k37_mlbage"] = [np.nan] * len(raw_df["ID"])
            raw_df["k38_nbbage"] = [np.nan] * len(raw_df["ID"])
            raw_df["k39_ukbage"] = [np.nan] * len(raw_df["ID"])

        if len(ped_sex) > 0:
            raw_df = raw_df.merge(ped_sex[['Participant Vehicle ID','ID','k40_fmped']], on=['ID','Participant Vehicle ID'], how= 'left')
            raw_df = raw_df.merge(ped_sex[['Participant Vehicle ID','ID','k41_mlped']], on=['ID','Participant Vehicle ID'], how= 'left')
            raw_df = raw_df.merge(ped_sex[['Participant Vehicle ID','ID','k42_nbped']], on=['ID','Participant Vehicle ID'], how= 'left')
            raw_df = raw_df.merge(ped_sex[['Participant Vehicle ID','ID','k43_ukped']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k40_fmped"] = [np.nan] * len(raw_df["ID"])
            raw_df["k41_mlped"] = [np.nan] * len(raw_df["ID"])
            raw_df["k42_nbped"] = [np.nan] * len(raw_df["ID"])
            raw_df["k43_ukped"] = [np.nan] * len(raw_df["ID"])

        if len(k44_dirvh) > 0:
            raw_df = raw_df.merge(k44_dirvh[['Participant Vehicle ID','ID','k44_dirvh']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k44_dirvh"] = [np.nan] * len(raw_df["ID"])

        if len(driver_res) > 0:
            raw_df = raw_df.merge(driver_res[['Participant Vehicle ID','ID','k45_lcdfl']], on=['ID','Participant Vehicle ID'], how= 'left')
            raw_df = raw_df.merge(driver_res[['Participant Vehicle ID','ID','k46_nlcdfl']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k45_lcdfl"] = [np.nan] * len(raw_df["ID"])
            raw_df["k46_nlcdfl"] = [np.nan] * len(raw_df["ID"])

        if len(k47_numvh) > 0:
            raw_df = raw_df.merge(k47_numvh[['Participant Vehicle ID','ID','k47_numvh']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k47_numvh"] = [np.nan] * len(raw_df["ID"])

        if len(k48_crshdt) > 0:
            raw_df = raw_df.merge(k48_crshdt[['Participant Vehicle ID','ID','k48_crshdt']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k48_crshdt"] = [np.nan] * len(raw_df["ID"])

        if len(k49_crshmo) > 0:
            raw_df = raw_df.merge(k49_crshmo[['Participant Vehicle ID','ID','k49_crshmo']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k49_crshmo"] = [np.nan] * len(raw_df["ID"])

        if len(k50_crshtm) > 0:
            raw_df = raw_df.merge(k50_crshtm[['Participant Vehicle ID','ID','k50_crshtm']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k50_crshtm"] = [np.nan] * len(raw_df["ID"])
        raw_df["k51_crshhr"] = raw_df["Crash Hour"]
        raw_df["k52_lat"] = raw_df["Latitude"]
        raw_df["k53_long"] = raw_df["Longitude"]

        if len(k54_coltyp) > 0:
            raw_df = raw_df.merge(k54_coltyp[['Participant Vehicle ID','ID','k54_coltyp']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k54_coltyp"] = [np.nan] * len(raw_df["ID"])

        if len(k55_vhclss) > 0:
            raw_df = raw_df.merge(k55_vhclss[['Participant Vehicle ID','ID','k55_vhclss']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k55_vhclss"] = [np.nan] * len(raw_df["ID"])

        raw_df["k56_sfcond"] = raw_df["Road Surface Condition"]
        raw_df["k57_wecond"] = raw_df["Weather Condition"]
        raw_df["k58_licond"] = raw_df["Light Condition"]
        if len(k59_siviol) > 0:
            raw_df = raw_df.merge(k59_siviol[['Participant Vehicle ID','ID','k59_siviol']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k59_siviol"] = [np.nan] * len(raw_df["ID"])

        raw_df["k60_crshsv"] = raw_df["Crash Severity"]

        if len(k61_crshpl) > 0:
            raw_df = raw_df.merge(k61_crshpl[['Participant Vehicle ID','ID','k61_crshpl']], on=['ID','Participant Vehicle ID'], how= 'left')
        else:
            raw_df["k61_crshpl"] = [np.nan] * len(raw_df["ID"])

        raw_df["k62_trnfl"] = k62_trnfl
        raw_df["k63_protfl"] = k63_protfl
        raw_df["k65_peddi"] = k65_peddi # Ped directly involved
        raw_df["k66_pcycdi"] = k66_pcycdi  # PedalCyclist directly involved
        raw_df = raw_df.merge(num_veh_dir, on=['ID','Participant Vehicle ID'], how= 'left')  # k67-75
        # raw_df.to_csv("ML_output_kai.csv", index=False)

    except Exception as err:
        st.text(err)
        # logging.error(err)
        sys.exit(1)
    return raw_df


def pivot_data(raw_df):
    try:
        # raw_df.to_csv(path_out + output_filename2,index=False)
        cols = raw_df.columns.tolist()
        cols.pop(0)
        cols.pop(0)
        raw_df["Temp ID"]=raw_df["ID"].astype(str)+raw_df["Participant Vehicle ID"]
        raw_df.drop_duplicates(subset=['Temp ID'],inplace=True)
        raw_df.reset_index(inplace=True)
        raw_df.drop(columns=["Temp ID", "index"], inplace=True)
        # ----------------- pivoting the data based on teh ID and Participant Vehicle ID
        raw_df = raw_df.pivot(index="ID", columns="Participant Vehicle ID", values=cols)
        # reorder the levels
        raw_df=raw_df.reorder_levels([1,0],axis=1)
        reorder_cols=[]
        # st.text(raw_df.columns)
        # gathering all columns with the keyword  Record Type
        for i in range (len(raw_df.columns)):
            if raw_df.columns[i][1] == "Record Type":
                reorder_cols.append(raw_df.columns[i][0])
        # st.text(reorder_cols)
        new_cols=raw_df.columns.reindex(reorder_cols,level=0)
        # st.text(new_cols)
        raw_df=raw_df.reindex(columns=new_cols[0])
        # # ---------------------
        # st.text(len(raw_df.columns))
        raw_df.dropna(how='all', axis=1, inplace=True)
        # st.text(len(raw_df.columns))
        raw_df.columns = raw_df.columns.map('|'.join).str.strip('|')  # merging the column headers
        raw_df.reset_index(inplace=True)
        raw_df = raw_df.rename(columns={'index': 'ID'})
        # raw_df.to_csv(path_out+output_filename1,index=False)
        return raw_df
    except Exception as err:
        st.text(err)
        # logging.error(err)
        sys.exit(1)


def generate_column_names(party, collision):
    try:
        def get_col_name(column_int): # convert index to alphabet column name in excel
            start_index = 0  # it can start either at 0 or at 1
            letter = ''
            while column_int > 25 + start_index:
                letter += chr(65 + int((column_int - start_index) / 26) - 1)
                column_int = column_int - (int((column_int - start_index) / 26)) * 26
            letter += chr(65 - start_index + (int(column_int)))
            return letter

        # column data used in the Column of Interest (COI) tab int eh visualizer
        input_tab = ["Collision", "Collision", "Collision", "Collision", "Collision", "Collision", "Collision", "Collision", "Collision",
                     "Collision", "Collision", "Collision", "Collision", "Collision", "Collision", "Collision", "Collision", "Collision",
                     "Collision", "Collision", "Party", "Party", "Party", "Collision"]

        output_tab = ["Severity", "Type", "Cause", "Event", "Temporal", "Temporal", "Temporal", "Temporal", "Road User", "Road User", "Road User",
                      "Road User", "Driver Cond", "Driver Condition", "Driver Condition", "Driver Condition", "Rdwy Cond", "Rdwy Cond",
                      "Rdwy Cond", "Location", "Age + Action", "Age + Action", "Age + Action", "Road User"]

        coi = ["Input Crash Severity Column", "Input Crash Type Column", "Input the first Crash Cause Column",
               "Input the first Crash Event Column", "Input Year Column", "Input Month Column", "Input Day of Week Column",
               "Input Hour Column", "Input Pedestrian Flag Column", "Input Pedalcyclist Flag Column", "Input Motorcyclist Flag Column",
               "Input Pedestrian Action Column", "Input Impairment Column", "Input Occupant Protection Flag Column",
               "Input Distracted Driving Flag Column", "Input Local Driver Flag Column", "Input Lighting Column", "Input Weather Column",
               "Input Road Surface Column", "Input Intersection Flag Column", "Input Party Type Column", "Input Age Column", "Input Party Action Column", "Pedalcyclist Collision Type"]

        col_name = ["General Crash Information|k01_sevr", "General Crash Information|k54_coltyp",
                    "General Crash Information|Crash Level Cause 1 Code", "General Crash Information|Crash Level Event 1 Code",
                    "General Crash Information|Crash Year", "General Crash Information|k49_crshmo", "General Crash Information|Week Day Code",
                    "General Crash Information|k51_crshhr", "General Crash Information|k65_peddi", "General Crash Information|k66_pcycdi",
                    "Vehicle 1|k12_motrfl", "Pedestrian/Pedalcyclist 1|Participant Action", "General Crash Information|k06_imprfl",
                    "General Crash Information|k63_protfl", "General Crash Information|k10_disdfl", "Vehicle 1 Participant 1|k45_lcdfl",
                    "General Crash Information|k58_licond", "General Crash Information|k57_wecond", "General Crash Information|k56_sfcond",
                    "General Crash Information|k11_intxfl", "Participant Type Code", "Age", "Participant Action", "General Crash Information|Collision Type"]

        col_id = []

        for i in range(len(input_tab)):
            if input_tab[i] == "Collision":
                try:
                    col_id.append(get_col_name(collision.columns.get_loc(col_name[i])))  # trying to get alphabet index based on column index
                except Exception as err:
                    col_id.append(None)
            elif input_tab[i] == "Party":
                try:
                    col_id.append(get_col_name(party.columns.get_loc(col_name[i])))
                except Exception as err:
                    col_id.append(None)
            else:
                pass

        df_col = pd.DataFrame()
        df_col["Input Tab"] = input_tab
        df_col["Output Tab"] = output_tab
        df_col["Column of Interest (COI)"] = coi
        df_col["Column Name"] = col_name
        df_col["Column ID"] = col_id
        return df_col

    except Exception as err:
        st.text(err)
        # logging.error(err)
        sys.exit(1)


def load_excel_from_url(url):
    return pd.ExcelFile(download_excel(url))


# Function to download the Excel file from URL
def download_excel(url):
    response = requests.get(url)
    response.raise_for_status()  # Ensure we notice bad responses
    return BytesIO(response.content)


# Function to convert DataFrame to Excel format, preserving formatting and formulas
def update_or_add_sheet_in_workbook(wb, sheet_name, df):

    ws = wb.create_sheet(title=sheet_name)

    for r_idx, row in df.iterrows():
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx + 2, column=c_idx + 1, value=value)  # assuming header is in the first row

    for idx, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=idx, value=col)

    return wb


def excel_table_export(df, df_pivot):
    viz_file = get_url("viz")

    try:
        workbook = load_workbook((viz_file))

        col_name_df = generate_column_names(df, df_pivot)  # generate column names
        df = df[df["Record Type"] == 3].reset_index()
        df.drop(columns="index", inplace=True)

        # df.to_csv(output_dir + "viz_party.csv", index= False)
        # df_pivot.to_csv(output_dir + "viz_col.csv", index=False)
        # col_name_df.to_csv(output_dir + "viz_coi.csv", index=False)

        try:
            # Remove existing placeholder worksheets
            for sheet in ['Collision', 'Party', 'COI']:
                if sheet in workbook.sheetnames:
                    del workbook[sheet]

            workbook = update_or_add_sheet_in_workbook(workbook, "Party", df)
            workbook = update_or_add_sheet_in_workbook(workbook, "Collision", df_pivot)
            workbook = update_or_add_sheet_in_workbook(workbook, "COI", col_name_df)

            # # Save the final output to BytesIO
            # final_output = BytesIO()
            # workbook.save(final_output)
            
            return workbook
            
        except Exception as err:
            st.text(err)
            # st.text("Worksheet(s) does not exist, check log file for next steps. However output file will be created....")
            # logging.info("There may be a duplicate in one or more Input tabs, delete the old tab(s) and rename the new one, to the same name as the old tab. However output file will be created....")
    
    except Exception as err:
        st.text(err)
        # logging.error(err)
        sys.exit(1)

def get_url(url_desc):
    if url_desc == "web_image":
        return 'brandfolder/Banner.png'
    elif url_desc == "dict":
        return "templates/Dictionary_OR.xlsx"
    elif url_desc == "viz":
        return "templates/Visualizer_OR.xlsx"
    else:
        pass


def download_all_files(col_df, col_name, party_df, party_name, viz_buffer, viz_name, zip_name):
    # Create in-memory buffers
    collision_buffer = io.BytesIO()
    party_buffer = io.BytesIO()

    collision_otput = col_df.to_csv(index=False).encode('utf-8')
    collision_buffer.write(collision_otput)
    
    party_otput = party_df.to_csv(index=False).encode('utf-8')
    party_buffer.write(party_otput)
    
    # Create a zip buffer
    zip_buffer = io.BytesIO()
    
    # Create a zip file in the buffer
    with zipfile.ZipFile(zip_buffer, 'w') as zf:
        zf.writestr(col_name, collision_buffer.getvalue())
        zf.writestr(party_name, party_buffer.getvalue())
        zf.writestr(viz_name, viz_buffer.getvalue())

     # Seek to the start of the zip buffer
    zip_buffer.seek(0)
    
    return zip_buffer


def handle_downloads (zip):
    if zip:
        # Create the download button for the zip file
        st.download_button(
            label="Download All Files",
            data=zip,
            file_name=output_filename + ".zip",
            mime='application/zip',
        )
        st.session_state.downloaded = True
    

# Main
if __name__ == '__main__':
    file_version = "Version 4"  # Kindly update this value after every version update
    project_name = "OR Crash Recoder Tool - " + file_version
    # # logging.basicConfig(level=# logging.DEBUG, filename="Log.log", filemode='a')
    st.image(get_url("web_image"), use_column_width=True)
    st.markdown(f"""<h1 style='text-align: center; font-size: 2em;'>{project_name}</h1>""", unsafe_allow_html=True)
    # logging.info("Crash Recoder Tool, an Innovation Kitchen Product developed by Kittelson and Associates, Inc. (KAI)")
    # logging.info("Contact: Azhagan (Azy) Avr  - aavr@kittelson.com")
    # logging.info(dt.datetime.now())
    st.text("\n  \n")
    if not st.session_state.downloaded:
        format_file = get_file_format("")
        if format_file:

            uploaded_files = get_uploaded_files(format_file)
        
            if uploaded_files:
                raw_data = concatenate_files(uploaded_files, format_file)

                # Getting dictionary and adding header tot eh raw data
                translation_df, raw_data_mod = get_dict_mod_raw_data(raw_data)
                output_filename = get_output_file_name()
                if output_filename:
                    
                    veh_code_seq = "0" # changing it to 0 as ODOT changed their format

                    output_filename1 = output_filename + "_Collision.csv"
                    output_filename2 = output_filename + "_Party.csv"
                    output_filename3 = output_filename + "_Visualizer.xlsx"

                    start_time = time.time()
                    st.text(str(round(time.time() - start_time, ndigits=2))+"s"+": Importing data translation table.....")
                    # logging.info(str(round(time.time() - start_time, ndigits=2)) + "s" + ": Importing data translation table.....")
                    translated_df = data_translation(raw_data_mod,translation_df, start_time)
                
                    st.text(str(round(time.time() - start_time, ndigits=2)) + "s" + ": Creating party level data.....")
                    # logging.info(str(round(time.time() - start_time, ndigits=2)) + "s" + ": Creating party level data.....")
                    pivot_col_id = participant_vehicle_id(translated_df, veh_code_seq)

                    st.text(str(round(time.time() - start_time, ndigits=2))+"s"+": Creating new variables.....")
                    # logging.info(str(round(time.time() - start_time, ndigits=2)) + "s" + ": Creating new variables.....")
                    new_var_df = add_kai_variables(pivot_col_id) # Outputfilename 2 df
                    new_var_df.to_csv(output_filename2, index= False)
                    st.text(str(round(time.time() - start_time, ndigits=2)) + "s" + ": Creating collision level data.....")
                    # logging.info(str(round(time.time() - start_time, ndigits=2)) + "s" + ": Creating collision level data.....")
                    pivot_df = pivot_data(new_var_df) # Output filename 1 df
                    pivot_df.to_csv(output_filename1, index=False)
                    st.text(str(round(time.time() - start_time, ndigits=2)) + "s" + ": Exporting data to visualizer.....")
                    # logging.info(str(round(time.time() - start_time, ndigits=2)) + "s" + ": Exporting data to visualizer.....")
                    
                    visualizer_buffer = excel_table_export(new_var_df, pivot_df) #visualizer file
                    visualizer_buffer.save(output_filename3)

                    # ....................................................................................................................
                    # logging.info(str(round((time.time() - start_time), ndigits=2)) + "s: Recoding complete.")
                    st.text(str(round((time.time() - start_time), ndigits=2)) + "s: Recoding complete.")

                    # zip = download_all_files(pivot_df, output_filename1, new_var_df, output_filename2, visualizer_buffer, output_filename3, output_filename)
                    
                    # if not st.session_state.downloaded:
                    #     handle_downloads(zip)

    else:
        time.sleep(2)
        st.markdown(f"""<h5 style='text-align: center; font-size: 1.5em;'>{"Your download will start shortly. Thank you for using the tool. Refresh the webpage to use the tool again."}</h5>""", unsafe_allow_html=True)
